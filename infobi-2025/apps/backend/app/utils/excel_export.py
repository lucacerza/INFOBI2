"""
Export Excel con mantenimento gerarchia Pivot
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from decimal import Decimal
from datetime import datetime, date
import io


def export_to_excel_with_pivot(data: List[Dict[str, Any]], pivot_config: Dict[str, Any]) -> bytes:
    """
    Esporta dati in Excel con gerarchia pivot e formattazione
    
    Args:
        data: Lista di dizionari con i dati
        pivot_config: Configurazione pivot {
            "row_groups": ["campo1", "campo2"],
            "columns": [...],
            "aggregations": {...}
        }
    
    Returns:
        bytes: File Excel in formato bytes
    """
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    
    if not data:
        # Foglio vuoto
        ws['A1'] = "Nessun dato disponibile"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # Estrai configurazione pivot
    row_groups = pivot_config.get("row_groups", [])
    
    # Se non ci sono gruppi, esporta come tabella flat
    if not row_groups:
        return _export_flat_table(data, wb, ws)
    
    # Altrimenti esporta con gerarchia
    return _export_with_hierarchy(data, row_groups, pivot_config, wb, ws)


def _export_flat_table(data: List[Dict[str, Any]], wb, ws) -> bytes:
    """Esporta tabella flat senza pivot"""
    
    if not data:
        ws['A1'] = "Nessun dato"
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    
    # Headers
    headers = list(data[0].keys())
    
    # Stili header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Dati
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header)
            value = _sanitize_cell_value(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Formattazione numeri
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
    
    # Auto-width colonne
    _auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _export_with_hierarchy(
    data: List[Dict[str, Any]], 
    row_groups: List[str],
    pivot_config: Dict[str, Any],
    wb, 
    ws
) -> bytes:
    """
    Esporta con gerarchia pivot
    Crea outline Excel con livelli collassabili
    """
    
    # Raggruppa dati per gerarchia
    grouped_data = _group_data_hierarchical(data, row_groups)
    
    # Headers
    all_columns = list(data[0].keys())
    
    # Stili
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    group_fills = [
        PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
        PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
        PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
    ]
    
    # Scrivi header
    for col_idx, header in enumerate(all_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Scrivi dati con gerarchia
    current_row = 2
    
    def write_group(group_data, level=0):
        nonlocal current_row
        
        for key, items in group_data.items():
            if isinstance(items, dict):
                # Sottogruppo
                # Scrivi riga gruppo
                cell = ws.cell(row=current_row, column=1, value=key)
                cell.font = Font(bold=True)
                if level < len(group_fills):
                    cell.fill = group_fills[level]
                
                # Imposta livello outline
                ws.row_dimensions[current_row].outline_level = level + 1
                
                current_row += 1
                
                # Ricorsione sui figli
                write_group(items, level + 1)
            else:
                # Dati foglia
                for row_data in items:
                    for col_idx, header in enumerate(all_columns, 1):
                        value = row_data.get(header)
                        value = _sanitize_cell_value(value)
                        
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right")
                    
                    # Imposta livello outline
                    ws.row_dimensions[current_row].outline_level = level + 1
                    
                    current_row += 1
    
    write_group(grouped_data)
    
    # Auto-width
    _auto_adjust_columns(ws)
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _group_data_hierarchical(data: List[Dict[str, Any]], group_fields: List[str]) -> Dict:
    """Raggruppa dati in struttura gerarchica"""
    
    if not group_fields:
        return data
    
    result = {}
    
    for row in data:
        current_level = result
        
        for i, field in enumerate(group_fields):
            key = str(row.get(field, ""))
            
            if i == len(group_fields) - 1:
                # Ultimo livello - aggiungi dati
                if key not in current_level:
                    current_level[key] = []
                current_level[key].append(row)
            else:
                # Livello intermedio
                if key not in current_level:
                    current_level[key] = {}
                current_level = current_level[key]
    
    return result


def _sanitize_cell_value(value: Any) -> Any:
    """Sanifica valore per Excel"""
    if value is None:
        return ""
    
    if isinstance(value, Decimal):
        return float(value)
    
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    
    return value


def _auto_adjust_columns(ws):
    """Auto-regola larghezza colonne"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
